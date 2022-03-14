# Web_server With the ability to accept multiple clients

#==========================================================================================
import logging
import threading
import time
import socket
import matplotlib.pyplot as plt
import numpy as np
import bs4 as bs
import openpyxl
import os
import datetime

#declaring global variables
line1 = []
thread = 0
device_map_dict = {'device_1_id':'device_1_id',
                    'device_2_id':'device_2_id',
                    'device_3_id':'device_3_id',
                    'device_4_id':'device_4_id',
                    'device_5_id':'device_5_id',
                    'device_6_id':'device_6_id',
                    'device_7_id':'device_7_id',
                    'device_8_id':'device_8_id',
                    'device_9_id':'device_9_id',
                    'device_10_id':'device_10_id'}
current_data_dict = {}
PatientID = 0

#==========================================================================================
#function for serving patient updated data
def patient_data(file):
    global current_data_dict, device_map_dict
    pt_sause = open('patientr/'+file).read()
    pt_soup =  bs.BeautifulSoup(pt_sause, 'html.parser')
    ID = file.split(".")[0]
    if ID in current_data_dict.keys():
        data_list = current_data_dict[ID]
        print(data_list)
    else:
        data_list = ['nodata', 'nodata', 'nodata', 'nodata']

    for i, row in enumerate(pt_soup.find_all('tr')):
        if i>0:
            col = row.find_all('td')
            col[1].clear()
            col[1].append(data_list[i-1])
            # print(col)
    pt_file = str(pt_soup)
    pt_file = "/\nHTTP/1.1 200 OK\n\n"+pt_file
    return pt_file

#==========================================================================================
# function for serving web_page

def web_page():
    http_response = open('patientr.html').read()
    http_response = "/\nHTTP/1.1 200 OK\n\n"+http_response
    return http_response

#==========================================================================================
# function for serving FetchData
def fetchData(file):
    ws = pd.read_excel(file, sheet_name = "patient data")
    # print(len(ws.index));
    rows = ws;
    if len(ws.index) < 10:
        rows = ws;
    else:
        rows = ws.iloc[-10:]
    arr = rows.to_numpy().tolist()
    print(arr)
    return arr

#==========================================================================================
# function for serving namechange

def namechange(folder):
    global device_map_dict
    sause = open('patientr.html').read()
    soup = bs.BeautifulSoup(sause, 'html.parser')
    dev_id = folder[2]
    i = int(dev_id.split("_")[1])
    button = soup.find_all('button')
    # print(type(button[i-1]))
    button[i-1].clear()
    button[i-1].append(folder[3])
    device_map_dict[dev_id] = folder[3]
    with open("patientr.html", "w", encoding = 'utf-8') as file:
        # prettify the soup object and convert it into a string  
        file.write(str(soup.prettify()))


#==========================================================================================

#function for Serving the client

def Client_thread(Client, address):
    print(threading.currentThread().getName(), 'Starting')
    global current_data_dict, thread, PatientID, device_map_dict
    request_data = Client.recv(1024)
    request_data = request_data.decode('utf-8')
    # print(request_data)
    code = request_data.split(" ", 2)

    if code[0] == "client" :
        PatientID = address
        device_ID = code[1]
        heading = ['SPO2', 'heart Rate', 'Respiration Rate', 'Temperature', 'Time']



        # Create a workbook and add a worksheet.
        # workbook = xlsxwriter.Workbook(device_ID+".xlsx")
        # worksheet = workbook.add_worksheet()
        if not os.path.isfile(device_map_dict[device_ID]+".xlsx"):
            wb = openpyxl.Workbook()
            wb['Sheet'].title = 'patient data'
            sh1 = wb.active
            for col in range(5):
                sh1.cell(1, col+1).value = heading[col]
            wb.save(device_map_dict[device_ID]+".xlsx")
        
        


        j = 0
        while True:
            data = Client.recv(2048)
            data = data.decode('utf-8')
            data_list = data.split(",")
            data_list.append(str(datetime.datetime.now()))
            current_data_dict[device_ID] = data_list
            print(data_list, address)
            if ('exit' in data):
                print("exit krra")
                Client.close()
                break
            if not os.path.isfile(device_map_dict[device_ID]+".xlsx"):
                wb = openpyxl.Workbook()
                wb['Sheet'].title = 'patient data'
                sh1 = wb.active
                for col in range(5):
                    sh1.cell(1, col+1).value = heading[col]
                wb.save(device_map_dict[device_ID]+".xlsx")

            wb = openpyxl.load_workbook(device_map_dict[device_ID]+".xlsx")
            sh1 = wb['patient data']
            max_row = sh1.max_row
            max_col = sh1.max_column
            # print('max row = %d , max col = %d\n',max_row, max_col)
            # print(data_list)
            for col in range(max_col):
                sh1.cell(row = max_row+1,column = col+1,value = data_list[col])
            wb.save(device_map_dict[device_ID]+".xlsx")

        Client.close()

    else:
        if code[1] == '/':
            Client.sendall(bytes(web_page(), "utf-8"))
            Client.close()

        else:
            folder = code[1].split("/", 3)
            print("FOLDER:", folder)
                # if folder[2] == 'fetchData':
                #     obj = [12, 13, 15, 18]
                #     data = json.dumps(obj)
    
                #     pt_file = "/\nHTTP/1.1 200 OK\n\n"
                #     Client.sendall(bytes(pt_file, "utf-8"))
                #     Client.sendall(bytes(data, "utf-8"))
                #     print("fetchData")
                #     Client.close()
                # elif folder[2] == 'fetchData/new':
                #     pass
            if folder[1] == 'patientr':
                file = folder[2]
                # print('file: 'file)
                if folder[3] == 'fetchData':   
                    obj = fetchData((file+'.xlsx'))
                    data = json.dumps(obj)
    
                    pt_file = "/\nHTTP/1.1 200 OK\n\n"
                    Client.sendall(bytes(pt_file, "utf-8"))
                    Client.sendall(bytes(data, "utf-8"))
                    print("fetchData")
                    Client.close()

                elif folder[3] == 'fetchNewData':
                    data = current_data_dict[file]
                    pt_file = "/\nHTTP/1.1 200 OK\n\n"
                    Client.sendall(bytes(pt_file, "utf-8"))
                    Client.sendall(bytes(data, "utf-8"))
                    print("fetchData")
                    Client.close()
                    
                else:
                    print("Data:", file)
                    Client.sendall(bytes(patient_data(file+'.html'), "utf-8"))
                    Client.close()
            elif folder[1] == 'namechange':
                namechange(folder)
                Client.close()
            else:
                Client.close()


    thread = thread-1
    print(threading.currentThread().getName(), 'Exiting')


#============================================================================================


#function for accepting clients

def Client_accept(ServerSocket):
    global thread
    print(threading.currentThread().getName(), 'Starting')
    while True:
        Client, address = ServerSocket.accept()
        print('Connected to: ' + address[0] + ':' + str(address[1]))
        # type(address[0]) = str, type(address[1]) = int
        client_thread = threading.Thread(name=address[0] + ':' + str(address[1]), target  = Client_thread, args = (Client, address, ))    #Separating the thread for serving
        thread = thread+1
        client_thread.start()                                                                                   #the client
    print(threading.currentThread().getName(), 'Exiting')

 #======================================================================================

#main function (Main Thread)

if __name__ == "__main__":
    ServerSocket = socket.socket()                                      #setting up the socket
    host = '10.12.0.65'                                                 #
    port = 4000                                                                 #
    ThreadCount = 0                                                             #
    ServerSocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    ServerSocket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)

    ServerSocket.bind((host, port))                                             #

    print('Waitiing for a Connection..')                                        #
    print("hostname " + socket.gethostbyname(host), port)                       #
    ServerSocket.listen(5)                                                      #
    print("Web server Started")                                         #server started
    acceptin_client = threading.Thread(name='accepting_client', target  = Client_accept, args = (ServerSocket, ))      #Making a Thread For accepting client
    acceptin_client.start()                                                                                         #starting the thread